#!/usr/bin/env python3
"""Normalize + dedup DCICA family-membership spreadsheets into a clean JSON
the importer (prisma/import-members.ts) can load.

The spreadsheets carry real PII and are NEVER committed; pass their paths and an
output path on the command line. Later files on the command line win ties on
dedup, so list newest last.

    python scripts/normalize-members.py OUT.json OLDEST.xlsx ... NEWEST.xlsx

Dedup key is the (lower-cased) email, falling back to a normalized name for the
email-less rows. On a collision the record with the later "Membership To" wins,
with missing phone / party-size / dates backfilled from the other. Rows with no
email get a stable placeholder (import-<namekey>@dcica.invalid) so the required,
unique email holds; nobody can accidentally mail a .invalid address.
"""
import sys, json, re, datetime, collections
import openpyxl

TODAY = datetime.datetime.now()


def norm_phone(p):
    if not p:
        return None
    d = re.sub(r"\D", "", str(p))
    d = d[1:] if len(d) == 11 and d[0] == "1" else d
    return d if len(d) == 10 else (str(p).strip() or None)


def name_key(n):
    return re.sub(r"[^a-z0-9]", "", n.lower())


def header_row(ws):
    for r in range(1, 8):
        if ws.cell(row=r, column=1).value == "Sno":
            return r
    return 2


def parse(src):
    ws = openpyxl.load_workbook(src, data_only=True)["Members"]
    hdr = header_row(ws)
    out = []
    for r in range(hdr + 1, ws.max_row + 1):
        v = [ws.cell(row=r, column=c).value for c in range(1, 17)]
        if all(x is None for x in v):
            continue
        _sno, name, fam, phone, email, _by, gs, mfrom, mto, *_ = v
        if not name:
            continue
        name = str(name).strip()
        email = (str(email).strip().lower() if email else None) or None
        is_life = (isinstance(mto, str) and "LIFE" in mto.upper()) or (
            isinstance(gs, str) and "LIFE" in gs.upper()
        )
        valid_to = None
        if isinstance(mto, datetime.datetime):
            valid_to = mto
        elif is_life:
            valid_to = datetime.datetime(2099, 12, 31)
        elif isinstance(gs, (int, float)) and 2000 < int(gs) < 2100:
            valid_to = datetime.datetime(int(gs), 12, 31)
        vf = mfrom if isinstance(mfrom, datetime.datetime) else None
        out.append(
            dict(
                name=name,
                email=email,
                phone=norm_phone(phone),
                partySize=int(fam) if isinstance(fam, (int, float)) and fam > 0 else None,
                isLife=is_life,
                validFrom=vf.date().isoformat() if vf else None,
                validTo=valid_to.date().isoformat() if valid_to else None,
            )
        )
    return out


def main():
    if len(sys.argv) < 3:
        sys.exit("usage: normalize-members.py OUT.json IN1.xlsx [IN2.xlsx ...]  (newest last)")
    out_path, srcs = sys.argv[1], sys.argv[2:]

    merged, collisions = {}, 0
    for src in srcs:  # newest last → wins ties
        for rr in parse(src):
            key = rr["email"] or "name:" + name_key(rr["name"])
            cur = merged.get(key)
            if cur is None:
                merged[key] = rr
                continue
            collisions += 1
            keep, drop = (rr, cur) if (rr["validTo"] or "") >= (cur["validTo"] or "") else (cur, rr)
            for f in ("phone", "partySize", "validFrom", "email"):
                if not keep.get(f) and drop.get(f):
                    keep[f] = drop[f]
            keep["isLife"] = cur["isLife"] or rr["isLife"]
            merged[key] = keep

    recs = list(merged.values())
    for r in recs:
        if not r["email"]:
            r["email"] = f"import-{name_key(r['name']) or 'unknown'}@dcica.invalid"
            r["emailSynthesized"] = True
        else:
            r["emailSynthesized"] = False
    # guarantee email uniqueness after synthesis
    seen = collections.Counter()
    for r in recs:
        seen[r["email"]] += 1
        if seen[r["email"]] > 1:
            local, domain = r["email"].split("@")
            r["email"] = f"{local}-{seen[r['email']]}@{domain}"

    json.dump(recs, open(out_path, "w"), indent=1)
    today = TODAY.date().isoformat()
    current = sum(1 for r in recs if r["isLife"] or (r["validTo"] and r["validTo"] >= today))
    print(f"{len(recs)} unique families ({collisions} dedup collisions resolved) -> {out_path}")
    print(f"  currently valid: {current}  (incl. {sum(1 for r in recs if r['isLife'])} LIFE)")
    print(f"  synthesized emails: {sum(1 for r in recs if r['emailSynthesized'])}")


if __name__ == "__main__":
    main()
